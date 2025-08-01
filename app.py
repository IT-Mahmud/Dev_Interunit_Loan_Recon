from flask import Flask, request, jsonify, render_template, send_from_directory
import os
import pandas as pd
from werkzeug.utils import secure_filename
from parser.tally_parser_interunit_loan_recon import parse_tally_file
from core import database
import threading
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Create upload folder
os.makedirs('uploads', exist_ok=True)

RECENT_UPLOADS_FILE = 'recent_uploads.txt'
RECENT_UPLOADS_LIMIT = 10
recent_uploads_lock = threading.Lock()

def record_recent_upload(filename):
    with recent_uploads_lock:
        try:
            if os.path.exists(RECENT_UPLOADS_FILE):
                with open(RECENT_UPLOADS_FILE, 'r', encoding='utf-8') as f:
                    uploads = [line.strip() for line in f if line.strip()]
            else:
                uploads = []
            # Remove if already present
            uploads = [f for f in uploads if f != filename]
            uploads.insert(0, filename)
            uploads = uploads[:RECENT_UPLOADS_LIMIT]
            with open(RECENT_UPLOADS_FILE, 'w', encoding='utf-8') as f:
                for f_name in uploads:
                    f.write(f_name + '\n')
        except Exception as e:
            print(f"Error recording recent upload: {e}")

@app.route('/api/recent-uploads', methods=['GET'])
def get_recent_uploads():
    try:
        if os.path.exists(RECENT_UPLOADS_FILE):
            with open(RECENT_UPLOADS_FILE, 'r', encoding='utf-8') as f:
                uploads = [line.strip() for line in f if line.strip()]
        else:
            uploads = []
        return jsonify({'recent_uploads': uploads})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clear-recent-uploads', methods=['POST'])
def clear_recent_uploads():
    try:
        if os.path.exists(RECENT_UPLOADS_FILE):
            with open(RECENT_UPLOADS_FILE, 'w', encoding='utf-8') as f:
                f.write('')
        return jsonify({'message': 'Recent uploads cleared.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def allowed_file(filename):
    """Check if file is Excel"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Upload and process file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        sheet_name = request.form.get('sheet_name', 'Sheet1')
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Please upload Excel files only'}), 400
        
        # Save file temporarily
        original_filename = file.filename
        filename = secure_filename(file.filename)
        filepath = os.path.join('uploads', filename)
        file.save(filepath)
        
        # Record recent upload (store original filename)
        record_recent_upload(original_filename)
        
        # Parse file
        df = parse_tally_file(filepath, sheet_name, original_filename)
        
        # Save to database
        success, error_msg = database.save_data(df)
        if success:
            os.remove(filepath)
            return jsonify({
                'message': 'File processed successfully',
                'rows_processed': len(df)
            })
        else:
            os.remove(filepath)
            return jsonify({'error': error_msg or 'Failed to save data'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload-pair', methods=['POST'])
def upload_file_pair():
    """Upload and process file pair"""
    try:
        if 'file1' not in request.files or 'file2' not in request.files:
            return jsonify({'error': 'Both files are required'}), 400
        
        file1 = request.files['file1']
        file2 = request.files['file2']
        sheet_name1 = request.form.get('sheet_name1', 'Sheet1')
        sheet_name2 = request.form.get('sheet_name2', 'Sheet1')
        
        if file1.filename == '' or file2.filename == '':
            return jsonify({'error': 'Both files must be selected'}), 400
        
        # Check if same file is uploaded twice
        if file1.filename == file2.filename:
            return jsonify({'error': 'Cannot upload the same file for both companies. Please select different files.'}), 400
        
        if not allowed_file(file1.filename) or not allowed_file(file2.filename):
            return jsonify({'error': 'Please upload Excel files only'}), 400
        
        # Generate unique pair ID
        import uuid
        from datetime import datetime
        pair_id = f"pair_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        
        total_rows = 0
        
        # Process first file
        original_filename1 = file1.filename
        filename1 = secure_filename(file1.filename)
        filepath1 = os.path.join('uploads', filename1)
        file1.save(filepath1)
        
        # Record recent upload
        record_recent_upload(original_filename1)
        
        # Parse first file
        df1 = parse_tally_file(filepath1, sheet_name1, original_filename1)
        # Add pair_id to first file
        df1['pair_id'] = pair_id
        total_rows += len(df1)
        
        # Save first file to database
        success1, error_msg1 = database.save_data(df1)
        if not success1:
            os.remove(filepath1)
            return jsonify({'error': error_msg1 or 'Failed to save first file'}), 400
        os.remove(filepath1)
        
        # Process second file
        original_filename2 = file2.filename
        filename2 = secure_filename(file2.filename)
        filepath2 = os.path.join('uploads', filename2)
        file2.save(filepath2)
        
        # Record recent upload
        record_recent_upload(original_filename2)
        
        # Parse second file
        df2 = parse_tally_file(filepath2, sheet_name2, original_filename2)
        # Add pair_id to second file
        df2['pair_id'] = pair_id
        total_rows += len(df2)
        
        # Save second file to database
        success2, error_msg2 = database.save_data(df2)
        if not success2:
            os.remove(filepath2)
            return jsonify({'error': error_msg2 or 'Failed to save second file'}), 400
        os.remove(filepath2)
        
        return jsonify({
            'message': 'File pair processed successfully',
            'rows_processed': total_rows,
            'pair_id': pair_id
        })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data', methods=['GET'])
def get_data():
    """Get all data"""
    try:
        data = database.get_data()
        # Get column order from database
        column_order = database.get_column_order()
        return jsonify({
            'data': data,
            'column_order': column_order
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/filters', methods=['GET'])
def get_filters():
    """Get filter options"""
    try:
        filters = database.get_filters()
        return jsonify(filters)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export', methods=['GET'])
def export_data():
    """Export filtered data to Excel"""
    try:
        filters = {}
        if request.args.get('lender'):
            filters['lender'] = request.args.get('lender')
        if request.args.get('borrower'):
            filters['borrower'] = request.args.get('borrower')
        if request.args.get('statement_month'):
            filters['statement_month'] = request.args.get('statement_month')
        if request.args.get('statement_year'):
            filters['statement_year'] = request.args.get('statement_year')
        
        data = database.get_data(filters)
        if not data:
            return jsonify({'error': 'No data found'}), 404
        
        df = pd.DataFrame(data)
        export_filename = f"tally_data_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        export_path = os.path.join('uploads', export_filename)
        
        df.to_excel(export_path, index=False)
        
        return send_from_directory('uploads', export_filename, as_attachment=True)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reconcile', methods=['POST'])
def reconcile_transactions():
    """Reconcile interunit transactions using new matching logic"""
    try:
        # Get all unmatched transactions
        data = database.get_unmatched_data()
        
        # Perform matching logic
        matches = database.find_matches(data)
        # Update database with matches
        database.update_matches(matches)
        
        return jsonify({
            'message': 'Reconciliation complete.',
            'matches_found': len(matches)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/matches', methods=['GET'])
def get_matches():
    """Get matched transactions"""
    try:
        matches = database.get_matched_data()
        return jsonify({'matches': matches})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pending-matches', methods=['GET'])
def get_pending_matches():
    """Get matches that need user confirmation"""
    try:
        matches = database.get_pending_matches()
        return jsonify({'matches': matches})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/confirmed-matches', methods=['GET'])
def get_confirmed_matches():
    """Get confirmed matches"""
    try:
        matches = database.get_confirmed_matches()
        return jsonify({'matches': matches})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/accept-match', methods=['POST'])
def accept_match():
    """Accept a match"""
    try:
        data = request.get_json()
        uid = data.get('uid')
        confirmed_by = data.get('confirmed_by', 'User')
        
        if not uid:
            return jsonify({'error': 'uid is required'}), 400
        
        success = database.update_match_status(uid, 'confirmed', confirmed_by)
        
        if success:
            return jsonify({'message': 'Match accepted successfully'})
        else:
            return jsonify({'error': 'Failed to accept match'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reject-match', methods=['POST'])
def reject_match():
    """Reject a match"""
    try:
        data = request.get_json()
        uid = data.get('uid')
        confirmed_by = data.get('confirmed_by', 'User')
        
        if not uid:
            return jsonify({'error': 'uid is required'}), 400
        
        success = database.update_match_status(uid, 'rejected', confirmed_by)
        
        if success:
            return jsonify({'message': 'Match rejected successfully'})
        else:
            return jsonify({'error': 'Failed to reject match'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-matches', methods=['GET'])
def download_matches():
    """Download matched transactions as Excel (matching table structure)"""
    try:
        matches = database.get_matched_data()
        if not matches:
            return jsonify({'error': 'No matched data found'}), 404
        df = pd.DataFrame(matches)
        # Get dynamic lender/borrower names from the first match
        lender_name = 'Lender'
        borrower_name = 'Borrower'
        if len(df) > 0:
            first_row = df.iloc[0]
            # Determine which is lender (Debit side) vs borrower (Credit side)
            if first_row.get('Debit') and first_row.get('Debit') > 0:
                # Main record is lender (Debit side)
                lender_name = first_row.get('lender', 'Lender')
                borrower_name = first_row.get('matched_lender', 'Borrower')
            elif first_row.get('matched_Debit') and first_row.get('matched_Debit') > 0:
                # Matched record is lender (Debit side)
                lender_name = first_row.get('matched_lender', 'Lender')
                borrower_name = first_row.get('lender', 'Borrower')
            else:
                # Fallback to original logic
                if first_row.get('lender'):
                    lender_name = first_row['lender']
                if first_row.get('matched_lender') and first_row['matched_lender'] != first_row.get('lender'):
                    borrower_name = first_row['matched_lender']
                elif first_row.get('borrower'):
                    borrower_name = first_row['borrower']
            
            # Build rows matching the table structure exactly
            export_rows = []
            for _, row in df.iterrows():
                # Determine which record is lender and which is borrower based on Debit/Credit values
                main_record_debit = float(row.get('Debit', 0) or 0)
                main_record_credit = float(row.get('Credit', 0) or 0)
                matched_record_debit = float(row.get('matched_Debit', 0) or 0)
                matched_record_credit = float(row.get('matched_Credit', 0) or 0)
                
                if main_record_debit > 0:
                    # Main record is Lender (Debit > 0)
                    lender_uid = row.get('uid')
                    lender_date = row.get('Date')
                    lender_particulars = row.get('Particulars')
                    lender_credit = row.get('Credit')
                    lender_debit = row.get('Debit')
                    lender_vch_type = row.get('Vch_Type')
                    
                    borrower_uid = row.get('matched_uid')
                    borrower_date = row.get('matched_date')
                    borrower_particulars = row.get('matched_particulars')
                    borrower_credit = row.get('matched_Credit')
                    borrower_debit = row.get('matched_Debit')
                    borrower_vch_type = row.get('matched_Vch_Type')
                elif matched_record_debit > 0:
                    # Matched record is Lender (Debit > 0)
                    lender_uid = row.get('matched_uid')
                    lender_date = row.get('matched_date')
                    lender_particulars = row.get('matched_particulars')
                    lender_credit = row.get('matched_Credit')
                    lender_debit = row.get('matched_Debit')
                    lender_vch_type = row.get('matched_Vch_Type')
                    
                    borrower_uid = row.get('uid')
                    borrower_date = row.get('Date')
                    borrower_particulars = row.get('Particulars')
                    borrower_credit = row.get('Credit')
                    borrower_debit = row.get('Debit')
                    borrower_vch_type = row.get('Vch_Type')
                else:
                    # Fallback: use the original logic based on lender_name
                    if row.get('lender') == lender_name:
                        lender_uid = row.get('uid')
                        lender_date = row.get('Date')
                        lender_particulars = row.get('Particulars')
                        lender_credit = row.get('Credit')
                        lender_debit = row.get('Debit')
                        lender_vch_type = row.get('Vch_Type')
                        
                        borrower_uid = row.get('matched_uid')
                        borrower_date = row.get('matched_date')
                        borrower_particulars = row.get('matched_particulars')
                        borrower_credit = row.get('matched_Credit')
                        borrower_debit = row.get('matched_Debit')
                        borrower_vch_type = row.get('matched_Vch_Type')
                    else:
                        borrower_uid = row.get('uid')
                        borrower_date = row.get('Date')
                        borrower_particulars = row.get('Particulars')
                        borrower_credit = row.get('Credit')
                        borrower_debit = row.get('Debit')
                        borrower_vch_type = row.get('Vch_Type')
                        
                        lender_uid = row.get('matched_uid')
                        lender_date = row.get('matched_date')
                        lender_particulars = row.get('matched_particulars')
                        lender_credit = row.get('matched_Credit')
                        lender_debit = row.get('matched_Debit')
                        lender_vch_type = row.get('matched_Vch_Type')
                
                # Determine roles based on Debit/Credit values
                if main_record_debit > 0:
                    lender_role = 'Lender'  # Main record has Debit > 0
                    borrower_role = 'Borrower'  # Matched record has Credit > 0
                elif matched_record_debit > 0:
                    lender_role = 'Lender'  # Matched record has Debit > 0
                    borrower_role = 'Borrower'  # Main record has Credit > 0
                else:
                    # Fallback: determine based on actual values
                    lender_role = 'Lender' if lender_debit and float(lender_debit) > 0 else 'Borrower'
                    borrower_role = 'Borrower' if borrower_credit and float(borrower_credit) > 0 else 'Lender'

                # Calculate matched amount
                matched_amount = max(
                    float(lender_debit or 0),
                    float(lender_credit or 0),
                    float(borrower_debit or 0),
                    float(borrower_credit or 0)
                )
                
                # Create row with proper column structure matching HTML table
                # The HTML table has 18 columns: 7 lender + 7 borrower + 4 match details
                row_data = {}
                
                # Lender section (columns 1-7) - match HTML table exactly
                row_data['Lender UID'] = lender_uid
                row_data['Lender Date'] = lender_date
                row_data['Lender Particulars'] = lender_particulars
                row_data['Lender Debit'] = lender_debit
                row_data['Lender Credit'] = lender_credit
                row_data['Lender Vch Type'] = lender_vch_type
                row_data['Lender Role'] = lender_role
                
                # Borrower section (columns 8-14) - match HTML table exactly
                row_data['Borrower UID'] = borrower_uid
                row_data['Borrower Date'] = borrower_date
                row_data['Borrower Particulars'] = borrower_particulars
                row_data['Borrower Debit'] = borrower_debit
                row_data['Borrower Credit'] = borrower_credit
                row_data['Borrower Vch Type'] = borrower_vch_type
                row_data['Borrower Role'] = borrower_role
                
                # Match details section (columns 15-18) - match HTML table exactly
                row_data['Confidence'] = f"{float(row.get('match_score', 0) * 100):.0f}%" if row.get('match_score') else 'N/A'
                row_data['Match Type'] = row.get('keywords') or 'Auto'
                row_data['Amount'] = matched_amount
                row_data['Actions'] = 'Pending'
                
                export_rows.append(row_data)
            
            export_df = pd.DataFrame(export_rows)
            export_filename = f"matched_transactions_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            export_path = os.path.join('uploads', export_filename)
            
            # Export with openpyxl for formatting
            with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
                # Creatfe a new DataFrame with section headers
                section_headers = []
                
                # Add section header row
                section_row = []
                section_row.extend(['LENDER (Debit > 0)'] * 7)  # 7 lender columns
                section_row.extend(['BORROWER (Credit > 0)'] * 7)  # 7 borrower columns  
                section_row.extend(['MATCH DETAILS'] * 4)  # 4 match detail columns
                section_headers.append(section_row)
                
                # Add column headers
                section_headers.append(export_df.columns.tolist())
                
                # Create header DataFrame
                header_df = pd.DataFrame(section_headers)
                
                # Combine headers with data
                final_df = pd.concat([header_df, export_df], ignore_index=True)
                final_df.to_excel(writer, index=False, sheet_name='Matched Transactions')
                
                # Get the worksheet for basic formatting
                worksheet = writer.sheets['Matched Transactions']
                
                # Set basic column widths
                for col_idx in range(1, len(export_df.columns) + 1):
                    col_letter = get_column_letter(col_idx)
                    column_name = export_df.columns[col_idx - 1]
                    
                    if 'Particulars' in column_name:
                        worksheet.column_dimensions[col_letter].width = 50
                    elif 'UID' in column_name:
                        worksheet.column_dimensions[col_letter].width = 20
                    elif 'Date' in column_name:
                        worksheet.column_dimensions[col_letter].width = 12
                    elif 'Debit' in column_name or 'Credit' in column_name or 'Amount' in column_name:
                        worksheet.column_dimensions[col_letter].width = 15
                    elif 'Confidence' in column_name:
                        worksheet.column_dimensions[col_letter].width = 12
                    elif 'Vch Type' in column_name:
                        worksheet.column_dimensions[col_letter].width = 15
                    elif 'Role' in column_name:
                        worksheet.column_dimensions[col_letter].width = 12
                    elif 'Match Type' in column_name:
                        worksheet.column_dimensions[col_letter].width = 15
                    elif 'Actions' in column_name:
                        worksheet.column_dimensions[col_letter].width = 12
                    else:
                        worksheet.column_dimensions[col_letter].width = 15
                
                # Format section headers (row 1)
                for col_idx in range(1, len(export_df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply background colors based on section
                    if col_idx <= 7:  # Lender section
                        cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
                    elif col_idx <= 14:  # Borrower section
                        cell.fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
                    else:  # Match details section
                        cell.fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
                
                # Format column headers (row 2)
                for col_idx in range(1, len(export_df.columns) + 1):
                    cell = worksheet.cell(row=2, column=col_idx)
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                
                # Freeze panes after headers
                worksheet.freeze_panes = "A3"
            
            return send_from_directory('uploads', export_filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/unmatched', methods=['GET'])
def get_unmatched():
    """Get unmatched transactions for display"""
    try:
        # Get query parameters for filtering
        lender_company = request.args.get('lender_company')
        borrower_company = request.args.get('borrower_company')
        month = request.args.get('month')
        year = request.args.get('year')
        
        if lender_company and borrower_company:
            # Get unmatched data for specific company pair
            data = database.get_unmatched_data_by_companies(lender_company, borrower_company, month, year)
        else:
            # Get all unmatched data
            data = database.get_unmatched_data()
        
        return jsonify({'unmatched': data})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-unmatched', methods=['GET'])
def download_unmatched():
    """Download unmatched transactions as Excel"""
    try:
        # Get query parameters for filtering
        lender_company = request.args.get('lender_company')
        borrower_company = request.args.get('borrower_company')
        month = request.args.get('month')
        year = request.args.get('year')
        
        if lender_company and borrower_company:
            # Get unmatched data for specific company pair
            data = database.get_unmatched_data_by_companies(lender_company, borrower_company, month, year)
        else:
            # Get all unmatched data
            data = database.get_unmatched_data()
        
        if not data:
            return jsonify({'error': 'No unmatched data found'}), 404
        
        df = pd.DataFrame(data)
        
        # Create a clean export DataFrame with key columns
        export_columns = [
            'uid', 'lender', 'borrower', 'statement_month', 'statement_year',
            'Date', 'Particulars', 'Vch_Type', 'Vch_No', 'Debit', 'Credit',
            'entered_by', 'input_date', 'role', 'original_filename'
        ]
        
        # Filter to only include columns that exist in the DataFrame
        available_columns = [col for col in export_columns if col in df.columns]
        export_df = df[available_columns].copy()
        
        # Rename columns for better readability
        column_mapping = {
            'uid': 'UID',
            'lender': 'Lender',
            'borrower': 'Borrower', 
            'statement_month': 'Statement Month',
            'statement_year': 'Statement Year',
            'Date': 'Date',
            'Particulars': 'Particulars',
            'Vch_Type': 'Voucher Type',
            'Vch_No': 'Voucher No',
            'Debit': 'Debit Amount',
            'Credit': 'Credit Amount',
            'entered_by': 'Entered By',
            'input_date': 'Input Date',
            'role': 'Role',
            'original_filename': 'Source File'
        }
        
        export_df = export_df.rename(columns=column_mapping)
        
        # Generate filename
        if lender_company and borrower_company:
            filename_suffix = f"{lender_company}_{borrower_company}"
            if month and year:
                filename_suffix += f"_{month}_{year}"
        else:
            filename_suffix = "all_data"
        
        export_filename = f"unmatched_transactions_{filename_suffix}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        export_path = os.path.join('uploads', export_filename)
        
        # Export with formatting
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Unmatched Transactions')
            
            # Get the worksheet for formatting
            worksheet = writer.sheets['Unmatched Transactions']
            
            # Set column widths
            for col_idx, column in enumerate(export_df.columns, 1):
                col_letter = get_column_letter(col_idx)
                
                if 'Particulars' in column:
                    worksheet.column_dimensions[col_letter].width = 50
                elif 'UID' in column:
                    worksheet.column_dimensions[col_letter].width = 25
                elif 'Date' in column:
                    worksheet.column_dimensions[col_letter].width = 12
                elif 'Debit' in column or 'Credit' in column:
                    worksheet.column_dimensions[col_letter].width = 15
                elif 'Source File' in column:
                    worksheet.column_dimensions[col_letter].width = 30
                else:
                    worksheet.column_dimensions[col_letter].width = 15
            
            # Format header row
            for col_idx in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            # Freeze header row
            worksheet.freeze_panes = "A2"
        
        return send_from_directory('uploads', export_filename, as_attachment=True)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/detected-pairs', methods=['GET'])
def get_detected_pairs():
    """Get automatically detected company pairs"""
    try:
        pairs = database.detect_company_pairs()
        return jsonify({'pairs': pairs})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/manual-pairs', methods=['GET'])
def get_manual_pairs():
    """Get manually defined company pairs"""
    try:
        pairs = database.get_manual_company_pairs()
        return jsonify({'pairs': pairs})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pairs', methods=['GET'])
def get_pairs():
    """Get all available pairs"""
    try:
        pairs = database.get_all_pair_ids()
        return jsonify({'pairs': pairs})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pair/<pair_id>/data', methods=['GET'])
def get_pair_data(pair_id):
    """Get all data for a specific pair"""
    try:
        data = database.get_data_by_pair_id(pair_id)
        return jsonify({'data': data, 'pair_id': pair_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pair/<pair_id>/unmatched', methods=['GET'])
def get_pair_unmatched_data(pair_id):
    """Get unmatched data for a specific pair"""
    try:
        data = database.get_unmatched_data_by_pair_id(pair_id)
        return jsonify({'unmatched': data, 'pair_id': pair_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reconcile-pair/<pair_id>', methods=['POST'])
def reconcile_pair(pair_id):
    """Reconcile transactions for a specific pair"""
    try:
        # Get unmatched transactions for this pair
        data = database.get_unmatched_data_by_pair_id(pair_id)
        
        # Perform matching logic
        matches = database.find_matches(data)
        # Update database with matches
        database.update_matches(matches)
        
        return jsonify({
            'message': f'Reconciliation complete for pair {pair_id}.',
            'matches_found': len(matches),
            'pair_id': pair_id
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 