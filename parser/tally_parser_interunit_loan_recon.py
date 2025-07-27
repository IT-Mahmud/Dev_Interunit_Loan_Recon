# tally_parser_interunit_loan_recon.py

import re
import pandas as pd
from openpyxl import load_workbook
from calendar import month_name
from typing import Tuple, Optional

def extract_statement_period(metadata: pd.DataFrame) -> Tuple[Tuple[str, str], str, Optional[int]]:
    period_pattern = re.compile(r'(\d{1,2}-[A-Za-z]{3}-\d{4})\s*to\s*(\d{1,2}-[A-Za-z]{3}-\d{4})')
    for i, row in metadata.iterrows():
        cell = str(row[0])
        match = period_pattern.search(cell)
        if match:
            return (match.group(1), match.group(2)), cell, i
    return ("", ""), "", None

def extract_lender(metadata: pd.DataFrame) -> Tuple[str, str, int]:
    unit_pattern = re.compile(r'Unit\s*:?[\s)]*([^)]+)')
    for i, row in metadata.iterrows():
        cell = str(row[0])
        match = unit_pattern.search(cell)
        if match:
            return match.group(1).strip(), cell, i
    return str(metadata.iloc[0, 0]).strip(), metadata.iloc[0, 0], 0

def extract_borrower(metadata: pd.DataFrame) -> Tuple[str, str, int]:
    pattern = re.compile(r'A/C-([\w\s&.()/-]+)')
    for i, row in metadata.iterrows():
        cell = str(row[0])
        match = pattern.search(cell)
        if match:
            borrower = match.group(1).strip()
            # Remove 'Unit' or 'unit' from the end (with or without period) and clean up
            borrower = re.sub(r'\s*[Uu]nit\.?\s*$', '', borrower).strip()
            # Replace 'Geo Textile' with 'GeoTex'
            if borrower == 'Geo Textile':
                borrower = 'GeoTex'
            return borrower, cell, i
    return "", "", None

def clean(val) -> str:
    cleaned = str(val).strip() if val is not None else ""
    # Remove _x000D_ characters (Carriage Return)
    cleaned = cleaned.replace('_x000D_', ' ')
    # Remove actual carriage return and line feed characters
    cleaned = cleaned.replace('\r', ' ')
    cleaned = cleaned.replace('\n', ' ')
    # Remove multiple spaces
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.strip()

def deduplicate_row(row, dup_map):
    res = row[:]
    for val, idxs in dup_map.items():
        found = False
        for i in idxs:
            if clean(res[i]) == val:
                if found:
                    res[i] = ""
                else:
                    found = True
    return res

def parse_tally_file(file_path: str, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    header_keywords = {"Date", "Particulars", "Vch Type", "Vch No.", "Debit", "Credit"}
    header_row_idx = next((i for i, r in enumerate(ws.iter_rows(values_only=True), 1)
                           if header_keywords.issubset({clean(c) for c in r})), None)
    if not header_row_idx:
        wb.close()
        raise ValueError("Header row not found.")

    metadata_rows = []
    for row in ws.iter_rows(min_row=1, max_row=header_row_idx-1, values_only=True):
        metadata_rows.append([clean(c) for c in row])
    metadata = pd.DataFrame(metadata_rows)

    (period_start, period_end), _, period_row = extract_statement_period(metadata)
    lender, _, lender_row = extract_lender(metadata)
    borrower, _, borrower_row = extract_borrower(metadata)

    ledger_date = ""
    ledger_year = ""
    if period_start and period_end:
        try:
            first_date = pd.to_datetime(period_start, format="%d-%b-%Y")
            last_date = pd.to_datetime(period_end, format="%d-%b-%Y")
            if first_date.month == last_date.month:
                ledger_date = month_name[first_date.month]
            if first_date.year == last_date.year:
                ledger_year = str(first_date.year)
        except Exception:
            pass

    for rng in list(ws.merged_cells.ranges):
        val = ws[rng.coord.split(":")[0]].value
        ws.unmerge_cells(str(rng))
        for row in ws[rng.coord]:
            for cell in row:
                cell.value = val

    headers = [clean(c.value) if c.value else f"Unnamed_{i+1}" for i, c in enumerate(ws[header_row_idx])]

    headers = ["dr_cr" if h == "Particulars" and i == headers.index("Particulars") else h for i, h in enumerate(headers)]
    particulars_index = headers.index("dr_cr") + 1
    if particulars_index < len(headers):
        headers[particulars_index] = "Particulars"

    num_cols = len(headers)

    collapsed_rows = []
    entered_by_list = []
    current_row = None
    last_entered_by = ""
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        cleaned = [clean(c) for c in row][:num_cols] + ["" for _ in range(num_cols - len(row))]
        entered_by_found = False
        for idx, cell in enumerate(cleaned):
            if "entered by :" in cell.lower():
                for next_idx in range(idx + 1, len(cleaned)):
                    if cleaned[next_idx]:
                        last_entered_by = cleaned[next_idx]
                        break
                else:
                    match = re.search(r"entered by\s*:\s*(.*)", cell, re.IGNORECASE)
                    if match:
                        last_entered_by = match.group(1).strip()
                entered_by_found = True
                break
        if entered_by_found:
            continue
        if (
            (not cleaned[headers.index("Date")] if "Date" in headers else True)
            and (not cleaned[headers.index("dr_cr")] if "dr_cr" in headers else True)
            and (cleaned[headers.index("Particulars")] if "Particulars" in headers else False)
            and current_row is not None
        ):
            idx = headers.index("Particulars")
            # Use space instead of newline since we're cleaning newlines
            current_row[idx] = (current_row[idx] + " " + cleaned[idx]).strip()
        else:
            if current_row is not None:
                collapsed_rows.append(current_row)
                entered_by_list.append(last_entered_by)
                last_entered_by = ""
            current_row = cleaned
    if current_row is not None:
        collapsed_rows.append(current_row)
        entered_by_list.append(last_entered_by)

    wb.close()
    dedup_map = {v: idxs for v, idxs in pd.Series(collapsed_rows[0]).groupby(
        lambda x: x).groups.items() if len(idxs) > 1}
    data_rows = [deduplicate_row(row, dedup_map) for row in collapsed_rows]

    if all(clean(v).replace('.', '', 1).replace(',', '', 1).isdigit() or clean(v) == "" for v in data_rows[-1]):
        data_rows.pop(-1)
        entered_by_list.pop(-1)

    df = pd.DataFrame(data_rows, columns=headers).dropna(axis=1, how='all')
    df = df.loc[:, (df != '').any(axis=0)]
    df = df.loc[:, ~df.columns.str.match(r'Unnamed_\d+')]

    df['entered_by'] = entered_by_list

    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(
            df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")

    if "Particulars" in df.columns:
        df = df[df["Particulars"].str.strip().str.lower() != "opening balance"]
        df = df[~df["Particulars"].str.strip().str.lower().str.startswith("closing balance")]

    def to_hex(val: str) -> str:
        try:
            return hex(int(float(val)))[2:]
        except Exception:
            return ""

    uids = []
    rownum = 1
    for i, row in df.iterrows():
        date_val = row.get("Date", "")
        credit_val = row.get("Credit", "")
        debit_val = row.get("Debit", "")
        balance_val = credit_val if (pd.notna(credit_val) and str(credit_val).strip() != "") else debit_val
        if pd.notna(date_val) and date_val != "":
            date_str = str(date_val).replace("-", "")
            hexdate = to_hex(date_str)
            try:
                hexbal = to_hex(round(float(str(balance_val).replace(",", ""))))
            except Exception:
                hexbal = ""
            # Add lender prefix to make tally_uid unique across files
            uid = f"{lender}_{hexdate}_{hexbal}_{rownum:06d}"
            uids.append(uid)
            rownum += 1
        else:
            uids.append("")
    df["tally_uid"] = uids
    cols = ["tally_uid", "lender", "borrower", "statement_month", "statement_year"] + \
        [c for c in df.columns if c not in ["tally_uid", "lender", "borrower", "statement_month", "statement_year"]]

    df["lender"] = lender
    df["borrower"] = borrower
    df["statement_month"] = ledger_date
    df["statement_year"] = ledger_year
    df = df[cols]

    if "Debit" in df.columns:
        df["Debit"] = df["Debit"].apply(lambda x: None if str(x).strip() == '' else x)
    if "Credit" in df.columns:
        df["Credit"] = df["Credit"].apply(lambda x: None if str(x).strip() == '' else x)

    new_column_names = {
        "Date": "Date",
        "Particulars": "Particulars",
        "Vch Type": "Vch_Type",
        "Vch No.": "Vch_No",
        "Debit": "Debit",
        "Credit": "Credit",
    }
    df = df.rename(columns=new_column_names)
    return df

if __name__ == "__main__":
    # Set these for IDE/Run Code button usage
    # input_file = "Input_Files/Interunit Steel.xlsx"
    # sheet_name = "Sheet7"

    input_file = "Input_Files/Interunit GeoTex.xlsx"
    sheet_name = "Sheet8"



    import sys
    if len(sys.argv) == 3:
        input_file = sys.argv[1]
        sheet_name = sys.argv[2]

    try:
        df = parse_tally_file(input_file, sheet_name)
        output_file = f"Parsed_{input_file.split('/')[-1]}"
        df.to_excel(output_file, index=False)
        print(f"Successfully parsed {input_file} and saved as {output_file}")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
